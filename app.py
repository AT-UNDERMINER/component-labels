"""
app.py — Flask web UI for the Component Label System (build step 9).

Thin presentation layer over the existing core/ modules — it owns no business
logic of its own:

    cache         read components + labels, persist edits
    label_builder render a component to label HTML (for the inline preview)
    pdf_renderer  render selected components to a downloadable A4 PDF
    config        sheet geometry, colour map, credential status, settings file

Pages (per CLAUDE.md > Web UI):
    /                       dashboard, filterable by type
    /component/<mpn>        detail: specs, images, datasheet, label preview
    /component/<mpn>/edit   edit: swap images, override specs, change type
    /print                  job builder -> downloadable PDF via render_sheet()
    /settings               geometry + colour editor, API credential status

Asset-serving helpers (so the inline label preview resolves its relative URLs):
    /label/<mpn>            raw build_label() HTML with a <base> injected
    /labelsrc/<path>        serves templates/ (label_styles.css)
    /output/<path>          serves output/ (pinout / package / qr images, PDFs)

Run with `python main.py web` (which calls main() below) or `python app.py`.
Binds 0.0.0.0 so it is reachable over the LAN / Tailscale; host/port/secret come
from the environment. WeasyPrint is only needed for /print; every other page
works without it, and /print degrades to a clean flash message if it is missing.
"""

from __future__ import annotations

import csv
import io
import json
import os
from datetime import datetime
from pathlib import Path

from flask import (
    Flask,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
    url_for,
)

import config
import main as cli  # the CLI orchestrator: add_component() + progress_sink()
                    # (aliased so it doesn't clash with this module's main())
from core import cache, label_builder, pdf_renderer
from core.datasheet import _safe_filename
from jobs import JobManager

app = Flask(
    __name__,
    template_folder="web/templates",
    static_folder="web/static",
)
# Flash messages need a session secret; a dev default is fine for a LAN tool.
app.secret_key = os.getenv("SECRET_KEY", "dev-component-labels")

# Background jobs for the dashboard's add / import features. In-memory, served
# by the threaded dev server; the browser polls /api/job/<id> for progress.
jobs = JobManager()


# ── Template helpers (available in every template via context processor) ───────

def _output_url(stored_path: str | Path | None) -> str | None:
    """Map a stored file path to its /output/<...> URL, or None if not under
    OUTPUT_DIR. Stored paths are project-root-relative (see cache._relativize),
    but this resolves them robustly however OUTPUT_DIR is configured."""
    if not stored_path:
        return None
    p = Path(stored_path)
    abs_p = p if p.is_absolute() else (config.PROJECT_ROOT / p)
    try:
        rel = abs_p.resolve().relative_to(config.OUTPUT_DIR.resolve())
    except (ValueError, OSError):
        return None
    return "/output/" + rel.as_posix()


@app.context_processor
def _inject_helpers() -> dict:
    return {
        "type_display_name": config.type_display_name,
        "type_colour": config.type_colour,
        "bar_text": label_builder.bar_text_colour,
        "component_types": config.COMPONENT_TYPES,
        "output_url": _output_url,
        "per_sheet": config.LABELS_PER_SHEET,
    }


def _preview_url(path: str | Path | None) -> str | None:
    """Map a stored path to a URL the label preview can load: /output/... for
    files under OUTPUT_DIR, /labelsrc/... for files under TEMPLATE_DIR (e.g. the
    package SVG library). None if the path is outside both trees."""
    if not path:
        return None
    p = Path(path)
    abs_p = (p if p.is_absolute() else config.PROJECT_ROOT / p).resolve()
    for base, prefix in ((config.OUTPUT_DIR, "/output/"), (config.TEMPLATE_DIR, "/labelsrc/")):
        try:
            return prefix + abs_p.relative_to(base.resolve()).as_posix()
        except (ValueError, OSError):
            continue
    return None


_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg"}


def _image_gallery(mpn: str) -> list[dict]:
    """Selectable images for the editor: the bundled package SVG library plus
    every image cached for this part (auto-extracted pinouts + uploads)."""
    items: list[dict] = []
    for svg in sorted(config.PACKAGE_DIR.glob("*.svg")):
        items.append({
            "path": svg.relative_to(config.PROJECT_ROOT).as_posix(),
            "url": "/labelsrc/" + svg.relative_to(config.TEMPLATE_DIR).as_posix(),
            "name": svg.stem,
            "kind": "package",
        })
    folder = config.IMAGE_DIR / _safe_filename(mpn)
    if folder.is_dir():
        for img in sorted(folder.rglob("*")):
            if (img.is_file() and img.suffix.lower() in _IMAGE_EXTS
                    and img.name != "qr.png"):
                rel = img.relative_to(config.PROJECT_ROOT).as_posix()
                items.append({"path": rel, "url": _output_url(rel),
                              "name": img.name, "kind": "image"})
    return items


def _spec_rows(record: dict, active_specs: list | None) -> list[dict]:
    """All API specs as picker rows, marking which are currently active and
    ordering active ones first (in their label order)."""
    active_vals = [a.get("value") for a in (active_specs or [])]
    order = {v: i for i, v in enumerate(active_vals)}
    rows: list[dict] = []
    for s in record.get("specs") or []:
        attr = s.get("attribute") or {}
        name = attr.get("name") or attr.get("shortname")
        val = s.get("displayValue")
        if not (name and val):
            continue
        rows.append({"name": str(name), "value": str(val),
                     "active": str(val) in active_vals})
    rows.sort(key=lambda r: (0, order.get(r["value"], 0)) if r["active"] else (1, 0))
    return rows


def _rebuild_label(mpn: str) -> None:
    """Re-render and store the label HTML after an edit (mirrors add step 6)."""
    cache.save_label(mpn, label_html=label_builder.build_label(mpn))


# ── Dashboard ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    type_filter = request.args.get("type") or None

    all_components = cache.list_components()
    counts: dict = {}
    for c in all_components:
        counts[c.get("component_type")] = counts.get(c.get("component_type"), 0) + 1

    if type_filter == "__none__":
        components = [c for c in all_components if not c.get("component_type")]
    elif type_filter:
        components = cache.list_components(type_filter)
    else:
        components = all_components

    # Split the (filtered) list into the dashboard's two sections. Anything not
    # explicitly 'needs_review' — including a NULL status from a pre-migration
    # row — counts as ready, so a component can never silently disappear.
    needs_review = [c for c in components if c.get("status") == "needs_review"]
    ready = [c for c in components if c.get("status") != "needs_review"]

    return render_template(
        "dashboard.html",
        components=components,
        needs_review=needs_review,
        ready=ready,
        counts=counts,
        total=len(all_components),
        type_filter=type_filter,
    )


# ── Background-job API (dashboard add + batch import) ──────────────────────────

def _is_checked(value: str | None) -> bool:
    """Interpret an HTML checkbox value as a bool."""
    return str(value).lower() in ("1", "true", "on", "yes")


@app.route("/api/lookup", methods=["POST"])
def api_lookup():
    """Feature A: start a single-part lookup in the background, return its id."""
    mpn = (request.form.get("mpn") or "").strip()
    if not mpn:
        return jsonify(error="Enter a part number."), 400

    ctype = (request.form.get("type") or "").strip() or None
    if ctype and ctype not in config.COMPONENT_TYPES:
        return jsonify(error=f"Unknown component type: {ctype}"), 400
    refresh = _is_checked(request.form.get("refresh"))

    job = jobs.create("lookup")
    job.set_result(mpn=mpn)
    jobs.start(job, lambda j: _run_lookup(j, mpn, ctype, refresh))
    return jsonify(job_id=job.id)


def _run_lookup(job, mpn: str, ctype: str | None, refresh: bool) -> None:
    """Worker: run the add pipeline, streaming its stages into the job log."""
    with cli.progress_sink(job.add_log):
        resolved = cli.add_component(mpn, component_type=ctype, refresh=refresh)
    job.set_result(mpn=resolved)


@app.route("/api/import", methods=["POST"])
def api_import():
    """Feature B: parse an uploaded .xlsx/.csv and start a batch import job."""
    file = request.files.get("file")
    if file is None or not (file.filename or "").strip():
        return jsonify(error="Choose a .xlsx or .csv file to import."), 400

    try:
        rows = _parse_import_file(file)
    except ValueError as exc:
        return jsonify(error=str(exc)), 400
    if not rows:
        return jsonify(error="No data rows found in the file."), 400

    refresh = _is_checked(request.form.get("refresh"))
    job = jobs.create("import")
    job.set_result(total=len(rows), done=0, succeeded=0, failed=0, skipped=0)
    jobs.start(job, lambda j: _run_import(j, rows, refresh))
    return jsonify(job_id=job.id, total=len(rows))


def _run_import(job, rows: list[dict], refresh: bool) -> None:
    """Worker: add each row sequentially, logging one line per part and keeping
    running succeeded/failed/skipped counts in the job result."""
    total = len(rows)
    succeeded = failed = skipped = 0
    seen: set[str] = set()

    for i, row in enumerate(rows, start=1):
        mpn = row["mpn"]
        ctype = row["type"] or None
        tag = f"[{i}/{total}]"

        if not mpn:
            skipped += 1
            job.add_log(f"{tag} (blank MPN) — skipped")
        elif mpn.lower() in seen:
            skipped += 1
            job.add_log(f"{tag} {mpn} — duplicate in file, skipped")
        elif ctype and ctype not in config.COMPONENT_TYPES:
            failed += 1
            job.add_log(f"{tag} {mpn} — unknown type '{ctype}', failed")
        else:
            seen.add(mpn.lower())
            job.add_log(f"{tag} {mpn} — adding...")
            try:
                resolved = cli.add_component(mpn, component_type=ctype, refresh=refresh)
                succeeded += 1
                job.add_log(f"{tag} {mpn} -> OK ({resolved})")
            except Exception as exc:  # one bad part must not abort the batch
                failed += 1
                job.add_log(f"{tag} {mpn} -> FAILED: {exc}")

        job.set_result(done=i, succeeded=succeeded, failed=failed, skipped=skipped)

    job.add_log(f"Done: {succeeded} succeeded, {failed} failed, {skipped} skipped.")


def _parse_import_file(file) -> list[dict]:
    """Read an uploaded .xlsx/.csv into [{"mpn", "type"}] rows.

    Requires an "mpn" header column; "type" is optional. Fully-blank rows are
    dropped. Raises ValueError (→ 400) for an unsupported or malformed file.
    """
    name = (file.filename or "").lower()
    if name.endswith(".csv"):
        text = file.read().decode("utf-8-sig", errors="replace")
        table = [list(r) for r in csv.reader(io.StringIO(text))]
    elif name.endswith(".xlsx"):
        from openpyxl import load_workbook
        wb = load_workbook(file, read_only=True, data_only=True)
        try:
            table = [list(r) for r in wb.active.iter_rows(values_only=True)]
        finally:
            wb.close()
    else:
        raise ValueError("Unsupported file type — upload a .xlsx or .csv file.")

    table = [r for r in table if any(c is not None and str(c).strip() for c in r)]
    if not table:
        raise ValueError("The file is empty.")

    header = [str(c).strip().lower() if c is not None else "" for c in table[0]]
    if "mpn" not in header:
        raise ValueError('Missing required "mpn" column in the header row.')
    mpn_i = header.index("mpn")
    type_i = header.index("type") if "type" in header else None

    def cell(raw: list, i: int | None) -> str:
        if i is None or i >= len(raw) or raw[i] is None:
            return ""
        return str(raw[i]).strip()

    return [{"mpn": cell(r, mpn_i), "type": cell(r, type_i)} for r in table[1:]]


@app.route("/api/import-template")
def import_template():
    """Serve a pre-filled .xlsx template (mpn, type) for the batch import."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Components"
    ws.append(["mpn", "type"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for example in [
        ("NE555P", "ic_opamp"),
        ("BC547", "bjt_transistor"),
        ("CF1/4W103JRC", "resistor"),
        ("B2B-XH-A(LF)(SN)", "connector"),
        ("WP7113ID", "led"),
    ]:
        ws.append(list(example))
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22

    # Dropdown of valid type keys on the type column (blank = auto-detect).
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(config.COMPONENT_TYPES) + '"',
        allow_blank=True,
    )
    dv.prompt = "Optional — leave blank to auto-detect the type."
    ws.add_data_validation(dv)
    dv.add("B2:B1000")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="component_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/api/job/<job_id>")
def api_job(job_id: str):
    """Poll a background job's status/log/result (used by the dashboard JS)."""
    job = jobs.get(job_id)
    if job is None:
        return jsonify(error="Unknown or expired job."), 404
    data = job.to_dict()
    # A finished lookup gets a link to its component page (needs request ctx).
    if data["kind"] == "lookup" and data["status"] == "done":
        mpn = data["result"].get("mpn")
        if mpn:
            data["result"]["url"] = url_for("component_detail", mpn=mpn)
    return jsonify(data)


# ── Component detail ─────────────────────────────────────────────────────────────

@app.route("/component/<mpn>")
def component_detail(mpn: str):
    record = cache.get_component(mpn)
    if record is None:
        abort(404)
    return render_template("detail.html", c=record, specs=record.get("specs") or [])


# ── Component edit ────────────────────────────────────────────────────────────────

@app.route("/component/<mpn>/edit", methods=["GET", "POST"])
def component_edit(mpn: str):
    """Visual drag-and-drop label editor.

    GET renders the two-panel editor (live preview + controls). POST is the
    AJAX save: it writes the editor's override state to the DB and rebuilds the
    stored label HTML, returning JSON {ok, url} for the JS to navigate to.
    """
    record = cache.get_component(mpn)
    if record is None:
        abort(404)

    if request.method == "POST":
        template = (request.form.get("template") or "").strip() or None
        headline = (request.form.get("headline") or "").strip()
        subline = (request.form.get("subline") or "").strip()
        try:
            specs_list = json.loads(request.form.get("specs") or "[]")
        except ValueError:
            specs_list = []

        # Persisted editor state. specs is always written (may be empty);
        # template/headline/subline only when set, so clearing a field reverts
        # to the computed default.
        overrides: dict = {"specs": specs_list[:3]}
        if template:
            overrides["template"] = template
        if headline:
            overrides["headline"] = headline
        if subline:
            overrides["subline"] = subline

        # "Approve" saves the edits AND clears the review flag: status -> ready,
        # review_reason -> None. A plain save passes status=None, which the
        # upsert preserves (so a needs_review part stays flagged until approved).
        approve = _is_checked(request.form.get("approve"))

        cache.upsert_component(
            mpn,
            overrides=overrides,
            pinout_image_path=(request.form.get("pinout_image") or "").strip() or None,
            package_image_path=(request.form.get("package_image") or "").strip() or None,
            status="ready" if approve else None,
            review_reason=None,
        )
        _rebuild_label(mpn)
        return jsonify(ok=True, url=url_for("component_detail", mpn=mpn))

    # GET: prefill controls from the current effective label context.
    context, template_name = label_builder.label_context(mpn)
    package = label_builder.package_image_path(record)
    return render_template(
        "editor.html",
        c=record,
        gallery=_image_gallery(mpn),
        spec_rows=_spec_rows(record, context.get("specs")),
        current_template=template_name,
        headline=context.get("headline") or "",
        subline=context.get("subline") or "",
        pinout_sel=(Path(record["pinout_image_path"]).as_posix()
                    if record.get("pinout_image_path") else ""),
        package_sel=(package.as_posix() if package else ""),
        template_options=[
            ("label_simple.html", "Simple"),
            ("label_complex.html", "Complex"),
            ("label_connector.html", "Connector"),
        ],
    )


@app.route("/api/upload-image/<mpn>", methods=["POST"])
def upload_image(mpn: str):
    """Save a user-uploaded image into the part's folder so the editor gallery
    can offer it. Returns its stored path + preview URL as JSON."""
    if cache.get_component(mpn) is None:
        return jsonify(error="Unknown component."), 404
    file = request.files.get("file")
    if file is None or not (file.filename or "").strip():
        return jsonify(error="No file selected."), 400
    ext = Path(file.filename).suffix.lower()
    if ext not in _IMAGE_EXTS:
        return jsonify(error="Use a PNG, JPG, GIF, WEBP or SVG image."), 400

    folder = config.IMAGE_DIR / _safe_filename(mpn) / "uploads"
    folder.mkdir(parents=True, exist_ok=True)
    stem = _safe_filename(Path(file.filename).stem)[:40] or "image"
    dest = folder / f"{stem}{ext}"
    n = 1
    while dest.exists():
        dest = folder / f"{stem}_{n}{ext}"
        n += 1
    file.save(str(dest))

    rel = dest.relative_to(config.PROJECT_ROOT).as_posix()
    return jsonify(path=rel, url=_output_url(rel), name=dest.name)


# ── Print job builder ────────────────────────────────────────────────────────────

@app.route("/print", methods=["GET", "POST"])
def print_builder():
    if request.method == "POST":
        mpns = request.form.getlist("mpns")
        if not mpns:
            flash("Select at least one component to print.", "error")
            return redirect(url_for("print_builder"))

        try:
            start = int(request.form.get("start", 1))
        except (TypeError, ValueError):
            flash("Start position must be a whole number.", "error")
            return redirect(url_for("print_builder"))
        if not (1 <= start <= config.LABELS_PER_SHEET):
            flash(f"Start position must be 1–{config.LABELS_PER_SHEET}.", "error")
            return redirect(url_for("print_builder"))

        missing = [m for m in mpns if not cache.component_exists(m)]
        if missing:
            flash("Not cached: " + ", ".join(missing), "error")
            return redirect(url_for("print_builder"))

        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out = config.LABEL_DIR / f"labels_{stamp}.pdf"
        try:
            pdf_renderer.render_sheet(mpns, out, start=start)
        except RuntimeError as exc:           # WeasyPrint missing, etc.
            flash(str(exc), "error")
            return redirect(url_for("print_builder"))
        except Exception as exc:              # any render failure -> clean message
            flash(f"PDF generation failed: {exc}", "error")
            return redirect(url_for("print_builder"))

        return send_file(
            out, as_attachment=True, download_name=out.name,
            mimetype="application/pdf",
        )

    return render_template("print.html", components=cache.list_components())


# ── Settings ──────────────────────────────────────────────────────────────────────

_GEOMETRY_FIELDS = [
    ("SHEET_WIDTH_MM", "Sheet width", "0.5", False),
    ("SHEET_HEIGHT_MM", "Sheet height", "0.5", False),
    ("LABEL_WIDTH_MM", "Label width", "0.1", False),
    ("LABEL_HEIGHT_MM", "Label height", "0.1", False),
    ("GRID_COLS", "Columns", "1", True),
    ("GRID_ROWS", "Rows", "1", True),
    ("MARGIN_TOP_MM", "Top margin", "0.1", False),
    ("MARGIN_LEFT_MM", "Left margin", "0.1", False),
    ("GAP_H_MM", "Horizontal gap", "0.1", False),
    ("GAP_V_MM", "Vertical gap", "0.1", False),
    ("BLEED_MM", "Bleed", "0.1", False),
    ("SAFE_ZONE_MM", "Safe zone inset", "0.1", False),
]


@app.route("/settings", methods=["GET", "POST"])
def settings():
    if request.method == "POST":
        geometry: dict = {}
        for name, _label, _step, is_int in _GEOMETRY_FIELDS:
            raw = (request.form.get(f"geo_{name}") or "").strip()
            if not raw:
                continue
            try:
                geometry[name] = int(raw) if is_int else float(raw)
            except ValueError:
                flash(f"Ignored invalid value for {name}: {raw!r}", "error")

        colours: dict = {
            key: request.form.get(f"col_{key}")
            for key in config.COMPONENT_TYPES
            if request.form.get(f"col_{key}")
        }

        try:
            config.SETTINGS_PATH.parent.mkdir(parents=True, exist_ok=True)
            config.SETTINGS_PATH.write_text(
                json.dumps({"geometry": geometry, "colours": colours}, indent=2),
                encoding="utf-8",
            )
        except OSError as exc:
            flash(f"Could not write settings: {exc}", "error")
            return redirect(url_for("settings"))

        flash("Settings saved to db/settings.json. Restart the server to apply.",
              "success")
        return redirect(url_for("settings"))

    geometry_fields = [
        {"name": name, "label": label, "step": step,
         "value": getattr(config, name)}
        for name, label, step, _is_int in _GEOMETRY_FIELDS
    ]
    api = {
        "client_id": bool(config.NEXAR_CLIENT_ID),
        "client_secret": bool(config.NEXAR_CLIENT_SECRET),
        "token_url": config.NEXAR_TOKEN_URL,
        "api_url": config.NEXAR_API_URL,
        "qr_host": config.QR_HOST,
    }
    return render_template("settings.html", geometry_fields=geometry_fields, api=api)


# ── Asset-serving routes (for the inline label preview + cached images) ─────────

@app.route("/label/<mpn>")
def label_html(mpn: str):
    """Raw single-label HTML for the preview iframe / editor live preview.

    With no query params this renders the component's saved label (saved editor
    overrides apply automatically via build_label). The editor passes its live,
    unsaved state as query params — template / headline / subline / specs (JSON)
    and pinout_image / package_image (stored paths) — which override the saved
    state so the preview updates in real time.

    URL resolution: a <base href="/labelsrc/"> makes the stylesheet link resolve
    via the labelsrc route; image overrides are converted to absolute
    /output/... or /labelsrc/... URLs (root-relative, so <base> ignores them).
    """
    record = cache.get_component(mpn)
    if record is None:
        abort(404)

    qr_path = config.IMAGE_DIR / _safe_filename(mpn) / "qr.png"

    ov: dict = {}
    for key in ("template", "headline", "subline"):
        if key in request.values:
            ov[key] = request.values.get(key)
    if "specs" in request.values:
        try:
            ov["specs"] = json.loads(request.values.get("specs") or "[]")
        except ValueError:
            pass

    def _img_override(param: str, default_path) -> str | None:
        """A passed image path wins (blank/"none" → no image); else the part's
        current default, both converted to a servable preview URL."""
        if param in request.values:
            chosen = (request.values.get(param) or "").strip()
            return _preview_url(chosen) if chosen and chosen.lower() != "none" else ""
        return _preview_url(default_path)

    ov["pinout_image"] = _img_override("pinout_image", record.get("pinout_image_path"))
    ov["package_image"] = _img_override("package_image", label_builder.package_image_path(record))
    ov["qr_image"] = _preview_url(qr_path)

    html = label_builder.build_label(mpn, overrides=ov)
    return html.replace("<head>", '<head><base href="/labelsrc/">', 1)


@app.route("/labelsrc/<path:filename>")
def labelsrc(filename: str):
    return send_from_directory(config.TEMPLATE_DIR, filename)


@app.route("/output/<path:filename>")
def output_files(filename: str):
    return send_from_directory(config.OUTPUT_DIR, filename)


# ── Entry point (invoked by `python main.py web`) ──────────────────────────────

def main() -> None:
    """Start the dev server. Host/port/debug come from the environment."""
    cache.init_db()
    host = os.getenv("FLASK_HOST", "0.0.0.0")
    port = int(os.getenv("FLASK_PORT") or os.getenv("PORT") or "5000")
    debug = os.getenv("FLASK_DEBUG", "").lower() in ("1", "true", "yes", "on")
    print(f"Component Labels web UI -> http://{host}:{port}  (Ctrl-C to stop)")
    # threaded=True so background add/import jobs run while the browser polls
    # /api/job/<id> for live progress.
    app.run(host=host, port=port, debug=debug, threaded=True)


if __name__ == "__main__":
    main()
